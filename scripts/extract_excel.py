#!/usr/bin/env python3
"""Extract data from Excel files (.xlsx).

Usage:
    python scripts/extract_excel.py <file.xlsx> [--sheets "Sheet1,Sheet2"] [--max-rows 200]

Examples:
    python scripts/extract_excel.py "data-room/Financial Model.xlsx"
    python scripts/extract_excel.py cap-table.xlsx --sheets "Cap Table,Summary"
    python scripts/extract_excel.py model.xlsx --max-rows 100

Requires: pip install openpyxl
"""

import sys
import io
import argparse

# Fix Windows Unicode output
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_cell(value) -> str:
    """Format a cell value for text output."""
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value) >= 1e6:
            return f"{value:,.0f}"
        elif abs(value) >= 1:
            return f"{value:,.2f}"
        else:
            return f"{value:.4f}"
    return str(value)


def main():
    parser = argparse.ArgumentParser(description="Extract data from Excel files")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--sheets", help="Comma-separated sheet names to extract (default: all)")
    parser.add_argument("--max-rows", type=int, default=300,
                        help="Max rows per sheet (default: 300)")
    parser.add_argument("--data-only", action="store_true", default=True,
                        help="Read formula results instead of formulas (default: True)")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.file, data_only=args.data_only, read_only=True)
    except Exception as e:
        print(f"ERROR: Cannot read Excel file: {e}", file=sys.stderr)
        sys.exit(1)

    all_sheets = wb.sheetnames
    print(f"Excel: {args.file} — {len(all_sheets)} sheets: {', '.join(all_sheets)}", file=sys.stderr)

    if args.sheets:
        target_sheets = [s.strip() for s in args.sheets.split(",")]
    else:
        target_sheets = all_sheets

    for sheet_name in target_sheets:
        if sheet_name not in all_sheets:
            print(f"\n=== Sheet: {sheet_name} === [NOT FOUND]")
            continue

        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")

        row_count = 0
        for row in ws.iter_rows(max_row=args.max_rows):
            cells = [format_cell(cell.value) for cell in row]
            # Skip completely empty rows
            if not any(cells):
                continue
            print("\t".join(cells))
            row_count += 1

        if row_count >= args.max_rows:
            print(f"[TRUNCATED at {args.max_rows} rows]")

        print(f"  {sheet_name}: {row_count} rows extracted", file=sys.stderr)

    wb.close()


if __name__ == "__main__":
    main()
